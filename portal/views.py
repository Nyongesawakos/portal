from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
import openpyxl
from django.contrib import messages
from .forms import StudentForm
from django.db.models import Sum
from django.http import JsonResponse
from .models import Student, CourseUnit, Result
from student.utils.grade import calculate_grade
from django.contrib.admin.views.decorators import staff_member_required

@login_required
def upload_students(request):
    excel = openpyxl.load_workbook(request.FILES['file'])
    sheet = excel.active

    headers = [cell.value for cell in sheet[1]]
    required = ["reg_no", "name", "email", "grade","fee_paid", "arrears", "assessment_no", "UPI", "Parent_contact"]

    if headers != required:
        return JsonResponse({"error": "Invalid student Excel format"})

    for row in sheet.iter_rows(min_row=2, values_only=True):
        Student.objects.update_or_create(
            reg_no=row[0],
            defaults={
                "name": row[1],
                "email": row[2],
                "grade": row[3],
                "fee_paid": row[4],
                "arrears": row[5],
                "assessment_no": row[6],
                "UPI": row[7],
                "Parent_contact": row[8],
            }
        )

    return HttpResponse({"uploaded successfully": "Students uploaded"})

@login_required
def upload_units(request):
    excel = openpyxl.load_workbook(request.FILES['file'])
    sheet = excel.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        CourseUnit.objects.update_or_create(
            code=row[0],
            defaults={
                "title": row[1],
                "credit_units": row[2]
            }
        )

    return JsonResponse({"success": "Course units uploaded"})

@login_required
def upload_results(request):
    excel = openpyxl.load_workbook(request.FILES['file'])
    sheet = excel.active

    for row in sheet.iter_rows(min_row=2, values_only=True):
        student = Student.objects.get(reg_no=row[0])
        unit = CourseUnit.objects.get(code=row[1])
        marks = row[2]

        Result.objects.update_or_create(
            student=student,
            unit=unit,
            defaults={
                "marks": marks,
                "grade": calculate_grade(marks)
            }
        )

    return JsonResponse({"success": "Results uploaded"})


# Create your views here.
@login_required
def add_student(request):
    form = StudentForm(request.POST or None)

    if request.method == "POST" and form.is_valid():
        form.save()
        return redirect('student_list')

    return render(request, 'add_student.html', {
        'form': form
    })
 # search Student
def student_list(request):
    query = request.GET.get('q')

    students = Student.objects.all()

    if query:
        students = students.filter(
            reg_no__icontains=query
        ) | students.filter(
            name__icontains=query
        ) | students.filter(
            email__icontains=query
        )

    return render(request, 'student_list.html', {
        'students': students,
        'query': query
    })

def announce(request):
    
    
    
  
    context ={}
    return render(request, 'announcements.html', context)

def profile(request, pk):
    std= Student.objects.get(id=pk)
    
    
    
  
    context ={'std': std}
    return render(request, 'profile.html', context)
def dashboard(request):
    std=Student.objects.all()
    stud= std.count()
    arrears= Student.objects.all().aggregate(total=Sum('arrears'))
    fee_paid= Student.objects.all().aggregate(total=Sum('fee_paid'))
    
    
    
  
    context ={'std': std, 'stud': stud, 'arrears': arrears['total'], 'fee_paid': fee_paid['total']}
    return render(request, 'dashboard.html', context)
def excel(request):
    
    
    
  
    context ={}
    return render(request, 'excel-upload.html', context)
def courses(request):
    
    
    
    
  
    context ={}
    return render(request, 'courses.html', context)


def updateStudent(request, pk):
    update = Student.objects.get(id=pk)
    form=StudentForm(instance=update)
    if request.method == 'POST':
        form=StudentForm(request.POST, instance=update)
        if form.is_valid():
            form.save()
            return redirect('student_list')
    context= {'update' : update, 'form' : form}
    return render(request, 'add_student.html', context)

def deleteStudent(request, pk):
    remove=Student.objects.get(id=pk)
    if request.method == 'POST':
        #name = Student.name 
        remove.delete()
        messages.success(request,  ' Student deleted successfully')
        return redirect('student_list')
    return render(request, 'delete.html', {'obj' :remove} )